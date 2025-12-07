"""
Find Source Chunks - Tim nguon chunk cho tung cau hoi
Dung fuzzy matching de tim chunk phu hop nhat voi cau tra loi
"""

import os
import json
import re
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def normalize_text(text: str) -> str:
    """Chuan hoa text de so sanh."""
    # Loai bo khoang trang thua, xuong dong
    text = ' '.join(text.split())
    # Loai bo so thu tu trang
    text = re.sub(r'\b\d{1,3}\b(?=\s|$)', '', text)
    return text.lower().strip()


def extract_keywords(text: str) -> set:
    """Trich xuat cac tu khoa quan trong tu text."""
    # Loai bo cac tu pho bien
    stopwords = {'la', 'cua', 'trong', 'va', 'cac', 'co', 'den', 'cho', 'voi',
                 'duoc', 'nhu', 'de', 'tu', 'mot', 'nhung', 'theo', 've',
                 'thi', 'khi', 'nay', 'do', 'se', 'da', 'hay', 'hoac', 'boi',
                 'là', 'của', 'trong', 'và', 'các', 'có', 'đến', 'cho', 'với',
                 'được', 'như', 'để', 'từ', 'một', 'những', 'theo', 'về',
                 'thì', 'khi', 'này', 'đó', 'sẽ', 'đã', 'hay', 'hoặc', 'bởi'}

    words = re.findall(r'\b\w+\b', text.lower())
    keywords = set(w for w in words if len(w) > 2 and w not in stopwords)
    return keywords


def calculate_similarity(text1: str, text2: str) -> float:
    """Tinh do tuong dong giua 2 text."""
    # Chuan hoa
    t1 = normalize_text(text1)
    t2 = normalize_text(text2)

    # Kiem tra tu khoa chung
    kw1 = extract_keywords(text1)
    kw2 = extract_keywords(text2)

    if not kw1 or not kw2:
        return 0.0

    # Do trung lap tu khoa
    common = kw1 & kw2
    keyword_score = len(common) / max(len(kw1), len(kw2)) if kw1 and kw2 else 0

    # Do tuong dong chuoi
    seq_score = SequenceMatcher(None, t1, t2).ratio()

    # Ket hop 2 diem
    return 0.6 * keyword_score + 0.4 * seq_score


def find_best_chunk(question: str, answer: str, chunks: dict, category: str) -> tuple:
    """
    Tim chunk phu hop nhat cho cau hoi.

    Returns:
        (chunk_name, chunk_content, similarity_score)
    """
    # Ket hop question va answer de tim kiem
    search_text = f"{question} {answer}"

    best_chunk = None
    best_content = ""
    best_score = 0.0

    for (cat, chunk_name), content in chunks.items():
        if cat != category:
            continue

        score = calculate_similarity(search_text, content)

        # Bonus: kiem tra cac cum tu quan trong co trong chunk khong
        answer_words = answer.split()
        if len(answer_words) >= 5:
            # Lay 5 tu lien tiep de kiem tra
            for i in range(len(answer_words) - 4):
                phrase = ' '.join(answer_words[i:i+5])
                if phrase.lower() in content.lower():
                    score += 0.3
                    break

        if score > best_score:
            best_score = score
            best_chunk = chunk_name
            best_content = content

    return best_chunk, best_content, best_score


def load_all_chunks(base_dir: str) -> dict:
    """Load tat ca chunk vao memory."""
    chunks = {}

    chunk_dirs = [
        ("culture", os.path.join(base_dir, "data_sources", "ban_sac_van_hoa_viet_nam", "structured_chunks_v2")),
        ("law", os.path.join(base_dir, "data_sources", "bai_giang_phap_luat_dai_cuong", "structured_chunks_v2")),
    ]

    for category, chunk_dir in chunk_dirs:
        if not os.path.exists(chunk_dir):
            continue
        for filename in sorted(os.listdir(chunk_dir)):
            if filename.endswith('.txt'):
                filepath = os.path.join(chunk_dir, filename)
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        content = f.read().strip()
                    chunks[(category, filename)] = content
                except Exception as e:
                    print(f"Warning: Could not read {filepath}: {e}")

    print(f"Loaded {len(chunks)} chunks")
    return chunks


def load_all_questions(base_dir: str) -> list:
    """Load tat ca cau hoi tu cac file JSON."""
    questions = []

    qa_dirs = [
        ("culture", os.path.join(base_dir, "data_question_answer", "ban_sac_van_hoa_viet_nam", "chunks")),
        ("law", os.path.join(base_dir, "data_question_answer", "bai_giang_phap_luat_dai_cuong", "chunks")),
    ]

    for category, qa_dir in qa_dirs:
        if not os.path.exists(qa_dir):
            continue
        for filename in sorted(os.listdir(qa_dir)):
            if filename.endswith('.json'):
                filepath = os.path.join(qa_dir, filename)
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        file_questions = json.load(f)
                    for q in file_questions:
                        q['_qa_file'] = filename  # Luu lai ten file JSON goc
                        q['category'] = category
                    questions.extend(file_questions)
                except Exception as e:
                    print(f"Warning: Could not load {filepath}: {e}")

    print(f"Loaded {len(questions)} questions")
    return questions


def load_existing_data(base_dir: str, questions: list) -> list:
    """Load topics, GPT answers, judge scores."""

    # Load topics
    topics_file = os.path.join(base_dir, "topics_progress.json")
    if os.path.exists(topics_file):
        try:
            with open(topics_file, 'r', encoding='utf-8') as f:
                topics = json.load(f)
            for q in questions:
                q_id = q.get("id", "")
                if q_id in topics:
                    q["predicted_topic"] = topics[q_id]
            print(f"  Loaded {len(topics)} topics")
        except Exception as e:
            print(f"  Warning: Error loading topics: {e}")

    # Load GPT answers
    for web_file in ["chatgpt_web_answers.json", "chatgpt_web_answers_law.json"]:
        web_path = os.path.join(base_dir, web_file)
        if os.path.exists(web_path):
            try:
                with open(web_path, 'r', encoding='utf-8') as f:
                    web_answers = json.load(f)
                matched = 0
                for q in questions:
                    q_id = q.get("id", "")
                    if q_id in web_answers:
                        answer_data = web_answers[q_id]
                        if isinstance(answer_data, dict):
                            q["gpt_answer"] = answer_data.get("gpt_web_answer", "")
                        else:
                            q["gpt_answer"] = str(answer_data)
                        matched += 1
                print(f"  Loaded {matched} GPT answers from {web_file}")
            except Exception as e:
                print(f"  Warning: Error loading {web_file}: {e}")

    # Load judge scores
    judge_file = os.path.join(base_dir, "judge_scores_progress.json")
    if os.path.exists(judge_file):
        try:
            with open(judge_file, 'r', encoding='utf-8') as f:
                judge_scores = json.load(f)
            for q in questions:
                q_id = q.get("id", "")
                if q_id in judge_scores:
                    q["judge_score"] = judge_scores[q_id].get("score")
                    q["judge_verdict"] = judge_scores[q_id].get("verdict", "")
                    q["judge_reason"] = judge_scores[q_id].get("reason", "")
            print(f"  Loaded {len(judge_scores)} judge scores")
        except Exception as e:
            print(f"  Warning: Error loading judge scores: {e}")

    return questions


def create_excel_with_sources(questions: list, output_path: str):
    """Tao file Excel voi cot Source Content."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation"

    # Headers
    headers = [
        "ID",
        "Source Chunk",
        "Source Content",
        "Match Score",
        "Category",
        "Predicted Topic",
        "Question",
        "Answer",
        "GPT_Answer",
        "GPT Score",
        "GPT Verdict",
        "GPT Reason",
        "text_based",
        "no_temporal",
        "relevant",
        "objective",
        "Total Score",
        "Notes"
    ]

    col_widths = [10, 20, 80, 10, 10, 18, 45, 50, 50, 10, 12, 40, 10, 10, 10, 10, 10, 30]

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap_align = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Set column widths
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Sort data
    sorted_data = sorted(questions, key=lambda x: (
        0 if x.get("gpt_answer") else 1,
        x.get("predicted_topic", "Khac"),
        x.get("id", "")
    ))

    # Topic colors
    topic_colors = {
        "Phap luat": "B4C6E7", "Pháp luật": "B4C6E7",
        "Van hoa": "F8CBAD", "Văn hóa": "F8CBAD",
        "Lich su": "FFE699", "Lịch sử": "FFE699",
        "Ton giao": "DDEBF7", "Tôn giáo": "DDEBF7",
        "Khac": "F2F2F2", "Khác": "F2F2F2"
    }

    # Write data rows
    for row_idx, item in enumerate(sorted_data, 2):
        col = 1

        # ID
        ws.cell(row=row_idx, column=col, value=item.get("id", "")).border = thin_border
        col += 1

        # Source Chunk
        ws.cell(row=row_idx, column=col, value=item.get("source_chunk", "")).border = thin_border
        col += 1

        # Source Content
        source_content = item.get("source_content", "")
        # Thay the ky tu xuong dong
        source_content = source_content.replace('\n', ' ').replace('\r', ' ')
        source_content = ' '.join(source_content.split())
        source_cell = ws.cell(row=row_idx, column=col, value=source_content)
        source_cell.alignment = wrap_align
        source_cell.border = thin_border
        col += 1

        # Match Score
        match_score = item.get("match_score", 0)
        score_cell = ws.cell(row=row_idx, column=col, value=f"{match_score:.2f}")
        score_cell.alignment = center_align
        score_cell.border = thin_border
        # Color by score
        if match_score >= 0.5:
            score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif match_score >= 0.3:
            score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        col += 1

        # Category
        ws.cell(row=row_idx, column=col, value=item.get("category", "")).border = thin_border
        col += 1

        # Predicted Topic
        topic = item.get("predicted_topic", "Khac")
        topic_cell = ws.cell(row=row_idx, column=col, value=topic)
        topic_cell.alignment = center_align
        topic_cell.border = thin_border
        if topic in topic_colors:
            topic_cell.fill = PatternFill(start_color=topic_colors[topic], end_color=topic_colors[topic], fill_type="solid")
        col += 1

        # Question
        q_cell = ws.cell(row=row_idx, column=col, value=item.get("question", ""))
        q_cell.alignment = wrap_align
        q_cell.border = thin_border
        col += 1

        # Answer
        a_cell = ws.cell(row=row_idx, column=col, value=item.get("answer", ""))
        a_cell.alignment = wrap_align
        a_cell.border = thin_border
        col += 1

        # GPT Answer
        gpt_cell = ws.cell(row=row_idx, column=col, value=item.get("gpt_answer", ""))
        gpt_cell.alignment = wrap_align
        gpt_cell.border = thin_border
        if item.get("gpt_answer"):
            gpt_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        col += 1

        # GPT Score
        judge_score = item.get("judge_score")
        score_cell = ws.cell(row=row_idx, column=col, value=judge_score if judge_score is not None else "")
        score_cell.alignment = center_align
        score_cell.border = thin_border
        if judge_score is not None:
            if judge_score >= 0.8:
                score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif judge_score >= 0.5:
                score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        col += 1

        # GPT Verdict
        ws.cell(row=row_idx, column=col, value=item.get("judge_verdict", "")).border = thin_border
        col += 1

        # GPT Reason
        reason_cell = ws.cell(row=row_idx, column=col, value=item.get("judge_reason", ""))
        reason_cell.alignment = wrap_align
        reason_cell.border = thin_border
        col += 1

        # Validation columns (empty)
        validation_start_col = col
        for _ in range(4):
            ws.cell(row=row_idx, column=col, value="").border = thin_border
            col += 1

        # Total Score formula
        validation_end_col = col - 1
        total_cell = ws.cell(
            row=row_idx,
            column=col,
            value=f"=SUM({get_column_letter(validation_start_col)}{row_idx}:{get_column_letter(validation_end_col)}{row_idx})"
        )
        total_cell.alignment = center_align
        total_cell.border = thin_border
        col += 1

        # Notes
        ws.cell(row=row_idx, column=col, value="").border = thin_border

    # Freeze header
    ws.freeze_panes = 'A2'

    # Auto filter
    last_col = get_column_letter(len(headers))
    last_row = len(sorted_data) + 1
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Save
    wb.save(output_path)
    print(f"\nExcel saved to: {output_path}")


def main():
    base_dir = r"D:\dichdata\vietnamese-culture-eval-2"
    output_path = os.path.join(base_dir, "vietnamese_benchmark_with_sources.xlsx")

    print("=" * 60)
    print("FIND SOURCE CHUNKS FOR QUESTIONS")
    print("=" * 60)

    # Load chunks
    print("\n[1/5] Loading chunks...")
    chunks = load_all_chunks(base_dir)

    # Load questions
    print("\n[2/5] Loading questions...")
    questions = load_all_questions(base_dir)

    # Load existing data
    print("\n[3/5] Loading existing data...")
    questions = load_existing_data(base_dir, questions)

    # Find source chunks
    print("\n[4/5] Finding source chunks for each question...")
    for i, q in enumerate(questions):
        if (i + 1) % 50 == 0:
            print(f"  Processing {i + 1}/{len(questions)}...")

        chunk_name, chunk_content, score = find_best_chunk(
            q.get("question", ""),
            q.get("answer", ""),
            chunks,
            q.get("category", "culture")
        )

        q["source_chunk"] = chunk_name or ""
        q["source_content"] = chunk_content or ""
        q["match_score"] = score

    # Statistics
    high_match = sum(1 for q in questions if q.get("match_score", 0) >= 0.5)
    medium_match = sum(1 for q in questions if 0.3 <= q.get("match_score", 0) < 0.5)
    low_match = sum(1 for q in questions if q.get("match_score", 0) < 0.3)

    print(f"\n  Match statistics:")
    print(f"    High (>=0.5): {high_match}")
    print(f"    Medium (0.3-0.5): {medium_match}")
    print(f"    Low (<0.3): {low_match}")

    # Create Excel
    print("\n[5/5] Creating Excel file...")
    create_excel_with_sources(questions, output_path)

    print("\n" + "=" * 60)
    print("COMPLETED!")
    print("=" * 60)


if __name__ == "__main__":
    main()
