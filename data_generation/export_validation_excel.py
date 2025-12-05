"""
Export Q&A Benchmark to Excel for Manual Validation

Creates an Excel file with columns for manual quality control:
- ID, Source, Category, Context, Question, Answer, GPT_Answer
- Validation columns for 4 criteria (1/0.5/0 scoring)
- Final score calculation

4 Validation Criteria:
1. text_based: Dựa hoàn toàn vào văn bản (Based entirely on text)
2. no_temporal: Không có thông tin thay đổi theo thời gian (No time-varying info)
3. relevant: Liên quan đến văn hóa hoặc pháp luật Việt Nam (Related to VN culture/law)
4. objective: Trung lập, khách quan, không ý kiến cá nhân (Neutral, objective)
"""

import os
import json
import openai
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Azure OpenAI Configuration
client = openai.AzureOpenAI(
    api_key="a6705b22532443ee8c0cfda232e57e06",
    azure_endpoint="https://vietgpt.openai.azure.com/",
    api_version="2024-02-15-preview"
)


def load_benchmark_data(json_paths: list) -> list:
    """
    Load Q&A data from JSON files.

    Args:
        json_paths: List of paths to JSON files

    Returns:
        List of Q&A dictionaries
    """
    all_data = []

    for json_path in json_paths:
        if os.path.exists(json_path):
            try:
                with open(json_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if isinstance(data, list):
                        all_data.extend(data)
                        print(f"  ✓ Loaded {len(data)} items from {os.path.basename(json_path)}")
            except Exception as e:
                print(f"  ✗ Error loading {json_path}: {str(e)}")
        else:
            print(f"  ⚠ File not found: {json_path}")

    return all_data


def load_benchmark_from_chunks(chunks_dir: str, source_type: str) -> list:
    """
    Load Q&A data from individual chunk JSON files.
    This preserves the chunk file information in the source field.

    Args:
        chunks_dir: Directory containing chunk JSON files
        source_type: "culture" or "law"

    Returns:
        List of Q&A dictionaries with proper source paths
    """
    all_data = []

    if not os.path.exists(chunks_dir):
        print(f"  ⚠ Chunks directory not found: {chunks_dir}")
        return all_data

    chunk_files = sorted([f for f in os.listdir(chunks_dir) if f.endswith('.json')])

    if not chunk_files:
        print(f"  ⚠ No chunk files found in {chunks_dir}")
        return all_data

    # Determine source path prefix based on type
    if source_type == "culture":
        source_prefix = "data_sources\\ban_sac_van_hoa_viet_nam\\structured_chunks_v2"
    else:
        source_prefix = "data_sources\\bai_giang_phap_luat_dai_cuong\\structured_chunks_v2"

    for chunk_file in chunk_files:
        file_path = os.path.join(chunks_dir, chunk_file)
        chunk_txt_name = chunk_file.replace('.json', '.txt')
        source_path = f"{source_prefix}\\{chunk_txt_name}"

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                questions = json.load(f)

            # Update source field for each question
            for q in questions:
                # Check if source is old format (just name) or already has path
                if q.get("source") and not q["source"].startswith("data_sources"):
                    q["source"] = source_path

            all_data.extend(questions)

        except Exception as e:
            print(f"  ✗ Error loading {chunk_file}: {str(e)}")

    print(f"  ✓ Loaded {len(all_data)} items from {len(chunk_files)} chunk files ({source_type})")
    return all_data


def get_gpt_answer(context: str, question: str) -> str:
    """
    Get GPT's answer for a question (without the source passage).
    This simulates how GPT would answer in web interface.

    Args:
        context: Context hint (e.g., "Sau đây là câu hỏi về văn hóa Việt Nam")
        question: The question to answer

    Returns:
        GPT's answer as string
    """
    try:
        # Combine context and question like in web interface
        user_message = f"{context}\n\n{question}" if context else question

        response = client.chat.completions.create(
            model="gpt-4o-2024-11-20",
            messages=[
                {
                    "role": "system",
                    "content": "Bạn là trợ lý AI có kiến thức về văn hóa và pháp luật Việt Nam. Trả lời ngắn gọn, chính xác, đi thẳng vào vấn đề. Không cần giải thích dài dòng."
                },
                {"role": "user", "content": user_message}
            ],
            temperature=0.3,
            max_tokens=500
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        return f"[ERROR: {str(e)}]"


def generate_gpt_answers(data: list, progress_file: str = None) -> list:
    """
    Generate GPT answers for all questions using API.
    Supports resuming from progress file.

    Args:
        data: List of Q&A dictionaries
        progress_file: Path to save/load progress

    Returns:
        List of Q&A dictionaries with gpt_answer field
    """
    # Load existing progress
    processed = {}
    if progress_file and os.path.exists(progress_file):
        try:
            with open(progress_file, 'r', encoding='utf-8') as f:
                processed = json.load(f)
            print(f"  Loaded {len(processed)} existing GPT answers")
        except:
            pass

    total = len(data)
    for idx, item in enumerate(data, 1):
        item_id = item.get("id", f"Q{idx}")

        # Skip if already processed
        if item_id in processed:
            item["gpt_answer"] = processed[item_id]
            continue

        print(f"  [{idx}/{total}] Getting GPT answer for {item_id}...")

        context = item.get("context", "")
        question = item.get("question", "")

        gpt_answer = get_gpt_answer(context, question)
        item["gpt_answer"] = gpt_answer
        processed[item_id] = gpt_answer

        # Save progress after each answer
        if progress_file:
            try:
                with open(progress_file, 'w', encoding='utf-8') as f:
                    json.dump(processed, f, ensure_ascii=False, indent=2)
            except:
                pass

    return data


def load_chatgpt_web_answers(web_answers_file: str, data: list) -> list:
    """
    Load câu trả lời từ ChatGPT web scraper và merge vào data.

    Args:
        web_answers_file: Path to chatgpt_web_answers.json
        data: List of Q&A dictionaries

    Returns:
        List of Q&A dictionaries with gpt_answer field
    """
    if not os.path.exists(web_answers_file):
        print(f"  ⚠ File not found: {web_answers_file}")
        return data

    try:
        with open(web_answers_file, 'r', encoding='utf-8') as f:
            web_answers = json.load(f)
        print(f"  ✓ Loaded {len(web_answers)} ChatGPT web answers")
    except Exception as e:
        print(f"  ✗ Error loading web answers: {e}")
        return data

    # Merge answers into data
    matched = 0
    for item in data:
        item_id = item.get("id", "")
        if item_id in web_answers:
            answer_data = web_answers[item_id]
            # Get the answer text
            if isinstance(answer_data, dict):
                item["gpt_answer"] = answer_data.get("gpt_web_answer", "")
            else:
                item["gpt_answer"] = str(answer_data)
            matched += 1

    print(f"  ✓ Matched {matched}/{len(data)} answers")
    return data


def create_validation_excel(data: list, output_path: str, include_gpt_answer: bool = True):
    """
    Create Excel file with validation columns for manual scoring.

    Args:
        data: List of Q&A dictionaries
        output_path: Path to output Excel file
        include_gpt_answer: Whether to include GPT_Answer column
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation"

    # Define headers based on whether GPT answer is included
    if include_gpt_answer:
        headers = [
            "ID",
            "Source",
            "Category",
            "Context",
            "Question",
            "Answer\n(Đáp án chuẩn)",
            "GPT_Answer\n(GPT trả lời)",
            "text_based\n(Dựa vào VB)",
            "no_temporal\n(Không TG)",
            "relevant\n(Liên quan)",
            "objective\n(Khách quan)",
            "Total Score",
            "Notes"
        ]
        col_widths = [10, 20, 10, 35, 45, 50, 50, 12, 12, 12, 12, 10, 30]
    else:
        headers = [
            "ID",
            "Source",
            "Category",
            "Context",
            "Question",
            "Answer",
            "text_based\n(Dựa vào VB)",
            "no_temporal\n(Không TG)",
            "relevant\n(Liên quan)",
            "objective\n(Khách quan)",
            "Total Score",
            "Notes"
        ]
        col_widths = [10, 20, 10, 35, 45, 55, 12, 12, 12, 12, 10, 30]

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap_align = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Set column widths
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Write data rows
    for row_idx, item in enumerate(data, 2):
        col = 1

        # ID
        ws.cell(row=row_idx, column=col, value=item.get("id", "")).border = thin_border
        col += 1

        # Source
        ws.cell(row=row_idx, column=col, value=item.get("source", "")).border = thin_border
        col += 1

        # Category
        ws.cell(row=row_idx, column=col, value=item.get("category", "")).border = thin_border
        col += 1

        # Context
        ctx_cell = ws.cell(row=row_idx, column=col, value=item.get("context", ""))
        ctx_cell.alignment = wrap_align
        ctx_cell.border = thin_border
        col += 1

        # Question
        q_cell = ws.cell(row=row_idx, column=col, value=item.get("question", ""))
        q_cell.alignment = wrap_align
        q_cell.border = thin_border
        col += 1

        # Answer (đáp án chuẩn)
        a_cell = ws.cell(row=row_idx, column=col, value=item.get("answer", ""))
        a_cell.alignment = wrap_align
        a_cell.border = thin_border
        col += 1

        # GPT Answer (nếu có)
        if include_gpt_answer:
            gpt_cell = ws.cell(row=row_idx, column=col, value=item.get("gpt_answer", ""))
            gpt_cell.alignment = wrap_align
            gpt_cell.border = thin_border
            # Highlight if GPT answer exists
            if item.get("gpt_answer"):
                gpt_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            col += 1

        # Validation columns (empty for manual input)
        validation_start_col = col
        for _ in range(4):  # 4 validation criteria
            cell = ws.cell(row=row_idx, column=col, value="")
            cell.alignment = center_align
            cell.border = thin_border
            col += 1

        # Total Score formula
        score_col = col
        validation_end_col = col - 1
        score_cell = ws.cell(
            row=row_idx,
            column=col,
            value=f"=SUM({get_column_letter(validation_start_col)}{row_idx}:{get_column_letter(validation_end_col)}{row_idx})"
        )
        score_cell.alignment = center_align
        score_cell.border = thin_border
        col += 1

        # Notes column
        notes_cell = ws.cell(row=row_idx, column=col, value="")
        notes_cell.alignment = wrap_align
        notes_cell.border = thin_border

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Add instruction sheet
    ws_guide = wb.create_sheet("Hướng dẫn")
    guide_content = [
        ["HƯỚNG DẪN ĐÁNH GIÁ Q&A BENCHMARK"],
        [""],
        ["Tiêu chí đánh giá (Scoring: 1 = Đạt, 0.5 = Gần đạt, 0 = Không đạt):"],
        [""],
        ["1. text_based (Dựa vào văn bản)"],
        ["   - 1: Câu hỏi và đáp án dựa hoàn toàn vào nội dung văn bản gốc"],
        ["   - 0.5: Phần lớn dựa vào văn bản, có một chút suy luận nhẹ"],
        ["   - 0: Suy luận ngoài văn bản hoặc thông tin không có trong nguồn"],
        [""],
        ["2. no_temporal (Không thời gian)"],
        ["   - 1: Không có thông tin thay đổi theo thời gian (GDP, dân số, năm cụ thể...)"],
        ["   - 0.5: Có đề cập nhưng không ảnh hưởng đến tính đúng đắn của câu hỏi"],
        ["   - 0: Có thông tin có thể lỗi thời (số liệu năm X, sự kiện cụ thể...)"],
        [""],
        ["3. relevant (Liên quan)"],
        ["   - 1: Hoàn toàn liên quan đến văn hóa hoặc pháp luật Việt Nam"],
        ["   - 0.5: Liên quan nhưng có phần chung chung hoặc không đặc thù VN"],
        ["   - 0: Không liên quan hoặc quá xa vời"],
        [""],
        ["4. objective (Khách quan)"],
        ["   - 1: Trung lập, khách quan, không có ý kiến cá nhân"],
        ["   - 0.5: Có một chút thiên vị nhưng không nghiêm trọng"],
        ["   - 0: Có ý kiến cá nhân, thiên vị rõ ràng"],
        [""],
        ["Tổng điểm tối đa: 4 điểm"],
        [""],
        ["Quy ước chất lượng:"],
        ["   - 4.0 điểm: Xuất sắc"],
        ["   - 3.0-3.5 điểm: Tốt"],
        ["   - 2.0-2.5 điểm: Trung bình (cần xem xét lại)"],
        ["   - < 2.0 điểm: Không đạt (cần loại bỏ hoặc sửa)"],
    ]

    for row_idx, row_content in enumerate(guide_content, 1):
        cell = ws_guide.cell(row=row_idx, column=1, value=row_content[0] if row_content else "")
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif "tiêu chí" in str(row_content[0]).lower() if row_content else False:
            cell.font = Font(bold=True, size=12)

    ws_guide.column_dimensions['A'].width = 80

    # Add summary sheet
    ws_summary = wb.create_sheet("Tổng kết")
    summary_headers = ["Metric", "Value"]
    ws_summary.append(summary_headers)

    total_items = len(data)
    ws_summary.append(["Tổng số câu hỏi", total_items])
    ws_summary.append(["Nguồn: Văn hóa", sum(1 for d in data if d.get("category") == "culture")])
    ws_summary.append(["Nguồn: Pháp luật", sum(1 for d in data if d.get("category") == "law")])
    ws_summary.append([""])
    ws_summary.append(["Điểm trung bình", f"=AVERAGE(Validation!J2:J{total_items + 1})"])
    ws_summary.append(["Số câu đạt (>=3)", f"=COUNTIF(Validation!J2:J{total_items + 1},\">=3\")"])
    ws_summary.append(["Số câu không đạt (<2)", f"=COUNTIF(Validation!J2:J{total_items + 1},\"<2\")"])

    # Style summary headers
    for cell in ws_summary[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20

    # Save workbook
    wb.save(output_path)
    print(f"\n✓ Excel file saved to: {output_path}")
    print(f"  - Total Q&A pairs: {len(data)}")
    print(f"  - Sheets: Validation, Hướng dẫn, Tổng kết")


def main():
    """Main function to export validation Excel."""
    import argparse

    parser = argparse.ArgumentParser(description='Export Q&A Benchmark to Validation Excel')
    parser.add_argument('--no-gpt', action='store_true',
                        help='Skip generating GPT answers (faster, no API calls)')
    parser.add_argument('--from-web', action='store_true',
                        help='Load GPT answers from ChatGPT web scraper (chatgpt_web_answers.json)')
    parser.add_argument('--web-file', type=str, default=None,
                        help='Path to ChatGPT web answers JSON file')
    parser.add_argument('--limit', type=int, default=None,
                        help='Limit number of questions to process (for testing)')
    parser.add_argument('--from-chunks', action='store_true',
                        help='Load from chunk files (preserves source file paths)')
    args = parser.parse_args()

    # Paths
    base_dir = r"D:/dichdata/vietnamese-culture-eval-2"
    qa_dir = os.path.join(base_dir, "data_question_answer")

    output_path = os.path.join(base_dir, "vietnamese_benchmark_validation.xlsx")
    progress_file = os.path.join(base_dir, "gpt_answers_progress.json")

    print("=" * 60)
    print("EXPORTING VIETNAMESE BENCHMARK TO VALIDATION EXCEL")
    print("=" * 60)

    # Load data
    print("\nLoading benchmark data...")

    if args.from_chunks:
        # Load from individual chunk files (preserves source paths)
        print("  (Loading from chunk files to preserve source paths)")
        data = []

        # Culture chunks
        culture_chunks_dir = os.path.join(qa_dir, "ban_sac_van_hoa_viet_nam", "chunks")
        data.extend(load_benchmark_from_chunks(culture_chunks_dir, "culture"))

        # Law chunks
        law_chunks_dir = os.path.join(qa_dir, "bai_giang_phap_luat_dai_cuong", "chunks")
        data.extend(load_benchmark_from_chunks(law_chunks_dir, "law"))
    else:
        # Load from merged benchmark files
        json_paths = [
            os.path.join(qa_dir, "ban_sac_van_hoa_viet_nam", "culture_benchmark.json"),
            os.path.join(qa_dir, "bai_giang_phap_luat_dai_cuong", "law_benchmark.json"),
        ]
        data = load_benchmark_data(json_paths)

    if not data:
        print("\n⚠ No data found. Please run generate_qa_benchmark.py first.")
        return

    # Apply limit if specified
    if args.limit:
        data = data[:args.limit]
        print(f"  Limited to {args.limit} questions for testing")

    include_gpt_answer = not args.no_gpt

    # Get GPT answers
    if args.from_web:
        # Load từ ChatGPT web scraper
        web_file = args.web_file or os.path.join(base_dir, "chatgpt_web_answers.json")
        print(f"\nLoading GPT answers from ChatGPT web scraper...")
        print(f"  File: {web_file}")
        data = load_chatgpt_web_answers(web_file, data)
        include_gpt_answer = True
    elif include_gpt_answer:
        # Generate qua API
        print(f"\nGenerating GPT answers via API for {len(data)} questions...")
        print("  (This may take a while. Progress is saved after each answer.)")
        print("  (You can stop and resume later.)")
        data = generate_gpt_answers(data, progress_file)
    else:
        print("\nSkipping GPT answer generation (--no-gpt flag)")

    # Create Excel
    print(f"\nCreating validation Excel with {len(data)} Q&A pairs...")
    create_validation_excel(data, output_path, include_gpt_answer)

    print("\n" + "=" * 60)
    print("EXPORT COMPLETE")
    print("=" * 60)
    print(f"\nOutput: {output_path}")

    if include_gpt_answer:
        print("\nColumns in Excel:")
        print("  - Source: Đường dẫn file chunk nguồn")
        print("  - Answer: Đáp án chuẩn (gen từ đoạn văn gốc)")
        print("  - GPT_Answer: Câu trả lời của GPT (không có đoạn văn)")
        print("\n→ So sánh 2 cột này để đánh giá GPT 'biết' hay 'không biết' câu trả lời")


if __name__ == "__main__":
    main()
