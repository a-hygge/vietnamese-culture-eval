"""
Export Final Excel - Xuat file Excel cuoi cung voi cot validation de trong
Khong can LLM, chi load du lieu va xuat Excel
"""

import os
import re
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# Cache for chunk contents: {chunk_name: content}
_chunk_content_cache = {}

# Cache for question_id -> chunk_name mapping
_question_chunk_map = {}


def build_question_chunk_mapping(base_dir: str):
    """
    Xay dung mapping tu question_id -> chunk_name.
    Doc cac file chunk_XXX.json trong thu muc data_question_answer/*/chunks/
    de biet cau hoi nao thuoc chunk nao.
    """
    global _question_chunk_map

    qa_chunk_dirs = [
        ("culture", os.path.join(base_dir, "data_question_answer", "ban_sac_van_hoa_viet_nam", "chunks")),
        ("law", os.path.join(base_dir, "data_question_answer", "bai_giang_phap_luat_dai_cuong", "chunks")),
    ]

    for category, chunk_dir in qa_chunk_dirs:
        if not os.path.exists(chunk_dir):
            continue
        for filename in os.listdir(chunk_dir):
            if filename.endswith('.json'):
                filepath = os.path.join(chunk_dir, filename)
                # chunk_001.json -> chunk_001.txt
                chunk_txt_name = filename.replace('.json', '.txt')
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        questions = json.load(f)
                    for q in questions:
                        q_id = q.get("id", "")
                        if q_id:
                            _question_chunk_map[q_id] = (category, chunk_txt_name)
                except:
                    pass

    print(f"  Built mapping for {len(_question_chunk_map)} questions")


def load_all_chunks(base_dir: str):
    """
    Load tat ca noi dung chunk vao cache.
    """
    global _chunk_content_cache

    # Load chunk content tu data_sources
    chunk_dirs = [
        ("culture", os.path.join(base_dir, "data_sources", "ban_sac_van_hoa_viet_nam", "structured_chunks_v2")),
        ("law", os.path.join(base_dir, "data_sources", "bai_giang_phap_luat_dai_cuong", "structured_chunks_v2")),
    ]

    loaded = 0
    for category, chunk_dir in chunk_dirs:
        if not os.path.exists(chunk_dir):
            continue
        for filename in os.listdir(chunk_dir):
            if filename.endswith('.txt'):
                filepath = os.path.join(chunk_dir, filename)
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        content = f.read().strip()
                    # Key = (category, filename)
                    _chunk_content_cache[(category, filename)] = content
                    loaded += 1
                except:
                    pass

    print(f"  Loaded {loaded} chunk files into cache")

    # Build question -> chunk mapping
    build_question_chunk_mapping(base_dir)


def get_chunk_content_for_item(item: dict, base_dir: str) -> str:
    """
    Lay noi dung chunk cho 1 item dua tren question ID.
    Thay the cac ky tu dac biet de tranh loi Excel.
    """
    q_id = item.get("id", "")

    if q_id in _question_chunk_map:
        category, chunk_name = _question_chunk_map[q_id]
        key = (category, chunk_name)
        if key in _chunk_content_cache:
            content = _chunk_content_cache[key]
            # Thay the ky tu xuong dong bang khoang trang de tranh loi Excel
            content = content.replace('\n', ' ').replace('\r', ' ')
            # Loai bo khoang trang thua
            content = ' '.join(content.split())
            return content

    return ""


def load_benchmark_data(base_dir: str) -> list:
    """Load benchmark data from JSON files."""
    all_data = []

    qa_dir = os.path.join(base_dir, "data_question_answer")

    # Culture benchmark
    culture_path = os.path.join(qa_dir, "ban_sac_van_hoa_viet_nam", "culture_benchmark.json")
    if os.path.exists(culture_path):
        with open(culture_path, 'r', encoding='utf-8') as f:
            culture_data = json.load(f)
        print(f"  Loaded {len(culture_data)} culture questions")
        all_data.extend(culture_data)

    # Law benchmark
    law_path = os.path.join(qa_dir, "bai_giang_phap_luat_dai_cuong", "law_benchmark.json")
    if os.path.exists(law_path):
        with open(law_path, 'r', encoding='utf-8') as f:
            law_data = json.load(f)
        print(f"  Loaded {len(law_data)} law questions")
        all_data.extend(law_data)

    return all_data


def load_existing_data(base_dir: str, data: list) -> list:
    """Load existing topics, GPT answers, and judge scores into data."""

    # Load topics
    topics_file = os.path.join(base_dir, "topics_progress.json")
    if os.path.exists(topics_file):
        try:
            with open(topics_file, 'r', encoding='utf-8') as f:
                topics = json.load(f)
            for item in data:
                item_id = item.get("id", "")
                if item_id in topics:
                    item["predicted_topic"] = topics[item_id]
            print(f"  Loaded {len(topics)} topic predictions")
        except Exception as e:
            print(f"  Warning: Error loading topics: {e}")

    # Load GPT web answers
    for web_file in ["chatgpt_web_answers.json", "chatgpt_web_answers_law.json"]:
        web_path = os.path.join(base_dir, web_file)
        if os.path.exists(web_path):
            try:
                with open(web_path, 'r', encoding='utf-8') as f:
                    web_answers = json.load(f)
                matched = 0
                for item in data:
                    item_id = item.get("id", "")
                    if item_id in web_answers:
                        answer_data = web_answers[item_id]
                        if isinstance(answer_data, dict):
                            item["gpt_answer"] = answer_data.get("gpt_web_answer", "")
                        else:
                            item["gpt_answer"] = str(answer_data)
                        matched += 1
                print(f"  Loaded {matched} GPT answers from {web_file}")
            except Exception as e:
                print(f"  Warning: Error loading {web_file}: {e}")

    # Load judge scores
    judge_file = os.path.join(base_dir, "judge_scores_progress.json")
    if os.path.exists(judge_file):
        try:
            with open(judge_file, 'r', encoding='utf-8') as f:
                judge_scores = json.load(f)
            for item in data:
                item_id = item.get("id", "")
                if item_id in judge_scores:
                    item["judge_score"] = judge_scores[item_id].get("score")
                    item["judge_verdict"] = judge_scores[item_id].get("verdict", "")
                    item["judge_reason"] = judge_scores[item_id].get("reason", "")
            print(f"  Loaded {len(judge_scores)} judge scores")
        except Exception as e:
            print(f"  Warning: Error loading judge scores: {e}")

    return data


def create_final_excel(data: list, output_path: str, base_dir: str):
    """
    Tao file Excel cuoi cung voi day du cot validation (de trong cho nguoi dung tu dien).
    Sort theo predicted_topic.
    Cot Source chua noi dung chunk.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation"

    # Headers - doi ten cot Source thanh "Source Content"
    headers = [
        "ID",
        "Source Content\n(Noi dung chunk)",
        "Category",
        "Predicted Topic",
        "Question",
        "Answer",
        "GPT_Answer",
        "GPT Score",
        "GPT Verdict",
        "GPT Reason",
        "text_based\n(Dua vao VB)",
        "no_temporal\n(Kg TG)",
        "relevant\n(Lien quan)",
        "objective\n(Khach quan)",
        "Total Score",
        "Notes"
    ]

    # Bo cot Context vi da co Source Content
    col_widths = [10, 80, 10, 18, 45, 50, 50, 10, 12, 40, 12, 12, 12, 12, 12, 30]

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap_align = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Set column widths
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Sort data:
    # 1. Cau co GPT answer len truoc (has_gpt_answer = 0 neu co, 1 neu khong)
    # 2. Theo predicted_topic
    # 3. Theo ID
    sorted_data = sorted(data, key=lambda x: (
        0 if x.get("gpt_answer") else 1,  # Co GPT answer len truoc
        x.get("predicted_topic", "Khac"),
        x.get("id", "")
    ))

    # Topic colors (Vietnamese without accents for compatibility)
    topic_colors = {
        "Phap luat": "B4C6E7",
        "Pháp luật": "B4C6E7",
        "Van hoa": "F8CBAD",
        "Văn hóa": "F8CBAD",
        "Lich su": "FFE699",
        "Lịch sử": "FFE699",
        "Kinh te": "C6EFCE",
        "Kinh tế": "C6EFCE",
        "Chinh tri": "D9E1F2",
        "Chính trị": "D9E1F2",
        "Xa hoi": "E2EFDA",
        "Xã hội": "E2EFDA",
        "Giao duc": "FCE4D6",
        "Giáo dục": "FCE4D6",
        "Ton giao": "DDEBF7",
        "Tôn giáo": "DDEBF7",
        "Nghe thuat": "FFF2CC",
        "Nghệ thuật": "FFF2CC",
        "Phong tuc tap quan": "E4DFEC",
        "Phong tục tập quán": "E4DFEC",
        "Dia ly": "D0CECE",
        "Địa lý": "D0CECE",
        "Khac": "F2F2F2",
        "Khác": "F2F2F2"
    }

    # Write data rows
    for row_idx, item in enumerate(sorted_data, 2):
        col = 1

        # ID
        ws.cell(row=row_idx, column=col, value=item.get("id", "")).border = thin_border
        col += 1

        # Source Content (noi dung chunk)
        chunk_content = get_chunk_content_for_item(item, base_dir)
        source_cell = ws.cell(row=row_idx, column=col, value=chunk_content)
        source_cell.alignment = wrap_align
        source_cell.border = thin_border
        col += 1

        # Category
        ws.cell(row=row_idx, column=col, value=item.get("category", "")).border = thin_border
        col += 1

        # Predicted Topic
        topic = item.get("predicted_topic", "Khac")
        topic_cell = ws.cell(row=row_idx, column=col, value=topic)
        topic_cell.alignment = center_align
        topic_cell.border = thin_border
        if topic in topic_colors:
            topic_cell.fill = PatternFill(start_color=topic_colors[topic], end_color=topic_colors[topic], fill_type="solid")
        col += 1

        # Question
        q_cell = ws.cell(row=row_idx, column=col, value=item.get("question", ""))
        q_cell.alignment = wrap_align
        q_cell.border = thin_border
        col += 1

        # Answer
        a_cell = ws.cell(row=row_idx, column=col, value=item.get("answer", ""))
        a_cell.alignment = wrap_align
        a_cell.border = thin_border
        col += 1

        # GPT Answer
        gpt_cell = ws.cell(row=row_idx, column=col, value=item.get("gpt_answer", ""))
        gpt_cell.alignment = wrap_align
        gpt_cell.border = thin_border
        if item.get("gpt_answer"):
            gpt_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        col += 1

        # GPT Score
        judge_score = item.get("judge_score")
        score_cell = ws.cell(row=row_idx, column=col, value=judge_score if judge_score is not None else "")
        score_cell.alignment = center_align
        score_cell.border = thin_border
        # Color by score
        if judge_score is not None:
            if judge_score >= 0.8:
                score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif judge_score >= 0.5:
                score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        col += 1

        # GPT Verdict
        verdict_cell = ws.cell(row=row_idx, column=col, value=item.get("judge_verdict", ""))
        verdict_cell.alignment = center_align
        verdict_cell.border = thin_border
        col += 1

        # GPT Reason
        reason_cell = ws.cell(row=row_idx, column=col, value=item.get("judge_reason", ""))
        reason_cell.alignment = wrap_align
        reason_cell.border = thin_border
        col += 1

        # Validation scores (4 columns - DE TRONG cho nguoi dung tu dien)
        validation_start_col = col
        for _ in range(4):
            v_cell = ws.cell(row=row_idx, column=col, value="")
            v_cell.alignment = center_align
            v_cell.border = thin_border
            col += 1

        # Total Score (formula)
        validation_end_col = col - 1
        total_cell = ws.cell(
            row=row_idx,
            column=col,
            value=f"=SUM({get_column_letter(validation_start_col)}{row_idx}:{get_column_letter(validation_end_col)}{row_idx})"
        )
        total_cell.alignment = center_align
        total_cell.border = thin_border
        col += 1

        # Notes (empty)
        notes_cell = ws.cell(row=row_idx, column=col, value="")
        notes_cell.alignment = wrap_align
        notes_cell.border = thin_border

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto filter
    last_col = get_column_letter(len(headers))
    last_row = len(sorted_data) + 1
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # =====================================================
    # Sheet: Huong dan
    # =====================================================
    ws_guide = wb.create_sheet("Huong dan")
    guide_content = [
        ["HUONG DAN DANH GIA Q&A BENCHMARK"],
        [""],
        ["=" * 60],
        ["TIEU CHI DANH GIA THU CONG (Scoring: 1 = Dat, 0.5 = Gan dat, 0 = Khong dat)"],
        ["=" * 60],
        [""],
        ["1. text_based (Dua vao van ban)"],
        ["   - 1: Cau hoi va dap an dua hoan toan vao noi dung van ban goc"],
        ["   - 0.5: Phan lon dua vao van ban, co mot chut suy luan nhe"],
        ["   - 0: Suy luan ngoai van ban hoac thong tin khong co trong nguon"],
        [""],
        ["2. no_temporal (Khong thay doi theo thoi gian)"],
        ["   - 1: Khong co thong tin thay doi theo thoi gian"],
        ["   - 0.5: Co de cap nhung khong anh huong den tinh dung dan cua cau hoi"],
        ["   - 0: Co thong tin co the loi thoi (so lieu nam X, su kien cu the...)"],
        [""],
        ["3. relevant (Lien quan den VN)"],
        ["   - 1: Hoan toan lien quan den van hoa hoac phap luat Viet Nam"],
        ["   - 0.5: Lien quan nhung co phan chung chung hoac khong dac thu VN"],
        ["   - 0: Khong lien quan hoac qua xa voi"],
        [""],
        ["4. objective (Khach quan)"],
        ["   - 1: Trung lap, khach quan, khong co y kien ca nhan"],
        ["   - 0.5: Co mot chut thien vi nhung khong nghiem trong"],
        ["   - 0: Co y kien ca nhan, thien vi ro rang"],
        [""],
        ["Tong diem toi da: 4 diem"],
        [""],
        ["Quy uoc chat luong:"],
        ["   - 4.0 diem: Xuat sac"],
        ["   - 3.0-3.5 diem: Tot"],
        ["   - 2.0-2.5 diem: Trung binh (can xem xet lai)"],
        ["   - < 2.0 diem: Khong dat (can loai bo hoac sua)"],
    ]

    for row_idx, row_content in enumerate(guide_content, 1):
        cell = ws_guide.cell(row=row_idx, column=1, value=row_content[0] if row_content else "")
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif "TIEU CHI" in str(row_content[0]).upper() if row_content else False:
            cell.font = Font(bold=True, size=12)

    ws_guide.column_dimensions['A'].width = 80

    # =====================================================
    # Sheet: Tong ket
    # =====================================================
    ws_summary = wb.create_sheet("Tong ket")

    # Calculate statistics
    total_items = len(data)

    # Count by category
    culture_count = sum(1 for d in data if d.get("category") == "culture")
    law_count = sum(1 for d in data if d.get("category") == "law")

    # Count by topic
    topic_counts = {}
    for item in data:
        topic = item.get("predicted_topic", "Khac")
        topic_counts[topic] = topic_counts.get(topic, 0) + 1

    # Judge score stats
    judge_scores = [d.get("judge_score") for d in data if d.get("judge_score") is not None]
    avg_judge = sum(judge_scores) / len(judge_scores) if judge_scores else 0

    # Write summary
    summary_data = [
        ["TONG KET BENCHMARK", ""],
        ["", ""],
        ["Tong so cau hoi", total_items],
        ["  - Van hoa (culture)", culture_count],
        ["  - Phap luat (law)", law_count],
        ["", ""],
        ["Co GPT answer", sum(1 for d in data if d.get("gpt_answer"))],
        ["Co Judge score", len(judge_scores)],
        ["GPT Judge Score TB", f"{avg_judge:.3f}" if judge_scores else "N/A"],
        ["", ""],
        ["PHAN BO CHU DE DU DOAN", ""],
    ]

    for topic, count in sorted(topic_counts.items(), key=lambda x: -x[1]):
        summary_data.append([f"  {topic}", f"{count} ({count/total_items*100:.1f}%)"])

    for row_idx, (col1, col2) in enumerate(summary_data, 1):
        ws_summary.cell(row=row_idx, column=1, value=col1)
        ws_summary.cell(row=row_idx, column=2, value=col2)
        if "TONG KET" in str(col1) or "PHAN BO" in str(col1):
            ws_summary.cell(row=row_idx, column=1).font = Font(bold=True)

    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20

    # Save
    wb.save(output_path)
    print(f"\n Excel saved to: {output_path}")
    print(f"  - Total: {total_items} questions")
    print(f"  - Sorted by: predicted_topic")
    print(f"  - Sheets: Validation, Huong dan, Tong ket")


def main():
    """Main function."""
    base_dir = r"D:\dichdata\vietnamese-culture-eval-2"
    output_path = os.path.join(base_dir, "vietnamese_benchmark_final_v5.xlsx")

    print("=" * 60)
    print("EXPORT FINAL EXCEL")
    print("=" * 60)

    # Load data
    print("\n[1/4] Loading benchmark data...")
    data = load_benchmark_data(base_dir)

    if not data:
        print("  No data found!")
        return

    # Load chunk contents
    print("\n[2/4] Loading chunk contents...")
    load_all_chunks(base_dir)

    # Load existing data
    print("\n[3/4] Loading existing data (topics, GPT answers, judge scores)...")
    data = load_existing_data(base_dir, data)

    # Create Excel
    print(f"\n[4/4] Creating final Excel file...")
    create_final_excel(data, output_path, base_dir)

    print("\n" + "=" * 60)
    print("COMPLETED!")
    print("=" * 60)
    print(f"\nOutput: {output_path}")
    print("\n4 cot validation (text_based, no_temporal, relevant, objective) de trong cho ban tu dien.")


if __name__ == "__main__":
    main()
