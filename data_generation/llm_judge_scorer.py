"""
LLM-as-a-Judge Scoring Script

Compares model responses against reference answers using GPT-4 as a judge.
Outputs: correct (1) / almost correct (0.5) / incorrect (0)

Output: Excel file with model responses and scores
"""

import os
import json
import openai
import pandas as pd
from typing import List, Dict, Literal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

# Import data quality checker
from data_quality_checker import (
    analyze_benchmark, analyze_benchmark_with_status, analyze_item,
    generate_quality_report, get_quality_status, classify_issues
)

# Azure OpenAI Configuration
client = openai.AzureOpenAI(
    api_key="a6705b22532443ee8c0cfda232e57e06",
    azure_endpoint="https://vietgpt.openai.azure.com/",
    api_version="2024-02-15-preview"
)

# ============================================================================
# PROMPTS
# ============================================================================

JUDGE_SYSTEM_PROMPT = """
Bạn là giám khảo đánh giá câu trả lời của mô hình AI về văn hóa và pháp luật Việt Nam.

## NHIỆM VỤ:
So sánh câu trả lời của mô hình với đáp án chuẩn và cho điểm.

## THANG ĐIỂM (0.0 - 1.0):
- 1.0 (excellent): Câu trả lời đúng hoàn toàn, đầy đủ và chính xác như đáp án chuẩn
- 0.9 (very_good): Đúng hoàn toàn về ý chính, chỉ khác biệt nhỏ về cách diễn đạt
- 0.8 (good): Đúng các điểm quan trọng, thiếu 1 chi tiết nhỏ không ảnh hưởng đến ý nghĩa
- 0.7 (fairly_good): Đúng phần lớn, thiếu 1-2 chi tiết phụ
- 0.6 (acceptable): Đúng ý chính nhưng thiếu một số thông tin bổ sung quan trọng
- 0.5 (partial): Đúng khoảng một nửa nội dung, thiếu nhiều thông tin quan trọng
- 0.4 (weak): Có một số ý đúng nhưng thiếu phần lớn nội dung chính
- 0.3 (poor): Chỉ đúng một phần nhỏ, phần lớn không chính xác hoặc thiếu
- 0.2 (very_poor): Hầu như không đúng, chỉ có liên quan mơ hồ đến đáp án
- 0.1 (almost_wrong): Gần như hoàn toàn sai, chỉ đề cập đúng chủ đề
- 0.0 (incorrect): Hoàn toàn sai, không liên quan, hoặc trả lời lạc đề

## NGUYÊN TẮC ĐÁNH GIÁ:
1. Đánh giá dựa trên NỘI DUNG và Ý NGHĨA, không phải cách diễn đạt
2. Chấp nhận các cách diễn đạt khác nhau nếu ý nghĩa tương đương
3. Ưu tiên độ chính xác của thông tin cốt lõi
4. Xem xét tính đầy đủ của câu trả lời so với đáp án chuẩn
5. Trừ điểm nặng nếu có thông tin SAI (không chỉ thiếu)
6. Câu trả lời dài hơn không có nghĩa là tốt hơn nếu thêm thông tin sai

## ĐỊNH DẠNG TRẢ LỜI:
{
  "score": <điểm từ 0.0 đến 1.0, bước 0.1>,
  "verdict": "excellent" | "very_good" | "good" | "fairly_good" | "acceptable" | "partial" | "weak" | "poor" | "very_poor" | "almost_wrong" | "incorrect",
  "reason": "<giải thích ngắn gọn lý do cho điểm>"
}

Chỉ trả về JSON, không giải thích thêm.
"""

JUDGE_USER_TEMPLATE = """
## CÂU HỎI:
{question}

## ĐÁP ÁN CHUẨN:
{correct_answer}

## CÂU TRẢ LỜI CỦA MÔ HÌNH:
{model_answer}

Hãy đánh giá câu trả lời của mô hình.
"""

# ============================================================================
# FUNCTIONS
# ============================================================================

def get_model_response(question: str, model_name: str = "gpt-4o-2024-11-20") -> str:
    """
    Get model response for a question.

    Args:
        question: The question to ask
        model_name: Model to use for response

    Returns:
        Model's answer as string
    """
    try:
        response = client.chat.completions.create(
            model=model_name,
            messages=[
                {"role": "system", "content": "Bạn là trợ lý AI có kiến thức về văn hóa và pháp luật Việt Nam. Trả lời ngắn gọn, chính xác."},
                {"role": "user", "content": question}
            ],
            temperature=0.3,
            max_tokens=500
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"  ✗ Error getting model response: {str(e)}")
        return f"[ERROR: {str(e)}]"


def judge_response(question: str, correct_answer: str, model_answer: str) -> Dict:
    """
    Use LLM as judge to score model response against reference answer.

    Args:
        question: The question asked
        correct_answer: The correct/reference answer
        model_answer: The model's response to evaluate

    Returns:
        Dictionary with score, verdict, and reason
    """
    try:
        user_prompt = JUDGE_USER_TEMPLATE.format(
            question=question,
            correct_answer=correct_answer,
            model_answer=model_answer
        )

        response = client.chat.completions.create(
            model="gpt-4o-2024-11-20",
            messages=[
                {"role": "system", "content": JUDGE_SYSTEM_PROMPT},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.1,
            max_tokens=500
        )

        result = response.choices[0].message.content.strip()

        # Parse JSON
        if "```json" in result:
            result = result.split("```json")[1].split("```")[0].strip()
        elif "```" in result:
            result = result.split("```")[1].split("```")[0].strip()

        return json.loads(result)

    except json.JSONDecodeError as e:
        print(f"  ✗ JSON parse error: {str(e)}")
        return {"score": 0, "verdict": "error", "reason": f"JSON parse error: {str(e)}"}
    except Exception as e:
        print(f"  ✗ Judge error: {str(e)}")
        return {"score": 0, "verdict": "error", "reason": f"Error: {str(e)}"}


def load_benchmark(json_path: str) -> List[Dict]:
    """Load benchmark data from JSON file."""
    if not os.path.exists(json_path):
        print(f"File not found: {json_path}")
        return []

    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading {json_path}: {str(e)}")
        return []


def evaluate_benchmark(benchmark_data: List[Dict], model_name: str = "gpt-4o-2024-11-20") -> List[Dict]:
    """
    Evaluate a benchmark dataset using LLM-as-a-Judge.

    Args:
        benchmark_data: List of Q&A pairs with id, question, answer
        model_name: Model to evaluate

    Returns:
        List of results with model responses and scores
    """
    results = []
    total = len(benchmark_data)

    # Pre-analyze data quality for all items
    print(f"\nAnalyzing data quality...")
    quality_results = analyze_benchmark_with_status(benchmark_data)
    critical_count = sum(1 for r in quality_results.values() if r["status"] == "CRITICAL")
    warning_count = sum(1 for r in quality_results.values() if r["status"] == "OK (warning)")
    print(f"  OK (warning): {warning_count} items (câu hỏi vẫn ổn)")
    print(f"  CRITICAL: {critical_count} items (cần xem lại)")

    print(f"\nEvaluating {total} questions with model: {model_name}")
    print("=" * 60)

    for idx, item in enumerate(benchmark_data, 1):
        question = item.get("question", "")
        correct_answer = item.get("answer", "")
        item_id = item.get("id", f"Q{idx}")

        print(f"\n[{idx}/{total}] {item_id}")

        # Get model response
        print(f"  → Getting model response...")
        model_answer = get_model_response(question, model_name)

        # Judge the response
        print(f"  → Judging response...")
        judgment = judge_response(question, correct_answer, model_answer)

        # Get data quality issues for this item
        quality_info = quality_results.get(item_id, {"issues": [], "status": "OK"})
        data_quality_status = quality_info["status"]
        item_issues = quality_info["issues"]
        data_quality_str = "; ".join(item_issues) if item_issues else "OK"

        result = {
            "id": item_id,
            "source": item.get("source", ""),
            "category": item.get("category", ""),
            "question": question,
            "correct_answer": correct_answer,
            "model_answer": model_answer,
            "score": judgment.get("score", 0),
            "verdict": judgment.get("verdict", "error"),
            "reason": judgment.get("reason", ""),
            "data_quality_status": data_quality_status,
            "data_quality_issues": data_quality_str
        }

        results.append(result)
        print(f"  ✓ Score: {result['score']} ({result['verdict']})")
        if data_quality_status == "CRITICAL":
            print(f"  ✗ CRITICAL: {len(item_issues)} issues")
        elif data_quality_status == "OK (warning)":
            print(f"  ✓ OK (warning)")

    return results


def export_results_to_excel(results: List[Dict], output_path: str, model_name: str):
    """
    Export evaluation results to Excel file.

    Args:
        results: List of evaluation results
        output_path: Path to output Excel file
        model_name: Name of the evaluated model
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluation Results"

    # Headers
    headers = [
        "ID",
        "Source",
        "Category",
        "Question",
        "Reference Answer",
        "Model Answer",
        "Score",
        "Verdict",
        "Reason",
        "Data Quality Issues"
    ]

    col_widths = [10, 25, 12, 45, 45, 45, 10, 15, 35, 50]

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    wrap_align = Alignment(vertical='top', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Score colors - gradient from green (1.0) to red (0.0)
    score_fills = {
        1.0: PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),  # Dark Green
        0.9: PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),  # Light Green
        0.8: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # Pale Green
        0.7: PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),  # Very Pale Green
        0.6: PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # Light Yellow
        0.5: PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # Yellow
        0.4: PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),  # Dark Yellow
        0.3: PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid"),  # Light Orange
        0.2: PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),  # Light Red
        0.1: PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid"),  # Red
        0.0: PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Dark Red
    }

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Set column widths
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Write data
    for row_idx, item in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=item.get("id", "")).border = thin_border
        ws.cell(row=row_idx, column=2, value=item.get("source", "")).border = thin_border
        ws.cell(row=row_idx, column=3, value=item.get("category", "")).border = thin_border

        # Question, reference, model answer
        for col, key in [(4, "question"), (5, "correct_answer"), (6, "model_answer")]:
            cell = ws.cell(row=row_idx, column=col, value=item.get(key, ""))
            cell.alignment = wrap_align
            cell.border = thin_border

        # Score with color
        score = item.get("score", 0)
        score_cell = ws.cell(row=row_idx, column=7, value=score)
        score_cell.alignment = center_align
        score_cell.border = thin_border
        if score in score_fills:
            score_cell.fill = score_fills[score]

        # Verdict
        verdict_cell = ws.cell(row=row_idx, column=8, value=item.get("verdict", ""))
        verdict_cell.alignment = center_align
        verdict_cell.border = thin_border

        # Reason
        reason_cell = ws.cell(row=row_idx, column=9, value=item.get("reason", ""))
        reason_cell.alignment = wrap_align
        reason_cell.border = thin_border

        # Data Quality Issues
        quality_status = item.get("data_quality_status", "OK")
        quality_issues = item.get("data_quality_issues", "")

        # Hiển thị status + issues
        if quality_status == "OK":
            display_text = "OK"
        elif quality_status == "OK (warning)":
            display_text = f"✓ OK (warning)\n{quality_issues}"
        else:
            display_text = f"✗ CRITICAL\n{quality_issues}"

        quality_cell = ws.cell(row=row_idx, column=10, value=display_text)
        quality_cell.alignment = wrap_align
        quality_cell.border = thin_border

        # Màu sắc theo status
        if quality_status == "OK":
            # Xanh lá đậm - hoàn toàn tốt
            quality_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif quality_status == "OK (warning)":
            # Xanh lá nhạt - vẫn ổn
            quality_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        else:
            # Cam - cần xem lại
            quality_cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")

    total = len(results)
    avg_score = sum(r.get("score", 0) for r in results) / total if total > 0 else 0

    # Count by score ranges
    score_counts = {}
    for score in [1.0, 0.9, 0.8, 0.7, 0.6, 0.5, 0.4, 0.3, 0.2, 0.1, 0.0]:
        score_counts[score] = sum(1 for r in results if r.get("score") == score)

    # Group scores into categories
    excellent_count = score_counts[1.0] + score_counts[0.9]  # 0.9-1.0
    good_count = score_counts[0.8] + score_counts[0.7]  # 0.7-0.8
    acceptable_count = score_counts[0.6] + score_counts[0.5]  # 0.5-0.6
    weak_count = score_counts[0.4] + score_counts[0.3]  # 0.3-0.4
    poor_count = score_counts[0.2] + score_counts[0.1] + score_counts[0.0]  # 0.0-0.2

    summary_data = [
        ["Vietnamese Benchmark Evaluation Summary", ""],
        ["", ""],
        ["Model Evaluated", model_name],
        ["Evaluation Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["Total Questions", total],
        ["Average Score", f"{avg_score:.3f}"],
        ["Total Points", f"{sum(r.get('score', 0) for r in results):.1f} / {total}"],
        ["", ""],
        ["Score Distribution:", ""],
        ["  Excellent (0.9-1.0)", f"{excellent_count} ({excellent_count/total*100:.1f}%)"],
        ["  Good (0.7-0.8)", f"{good_count} ({good_count/total*100:.1f}%)"],
        ["  Acceptable (0.5-0.6)", f"{acceptable_count} ({acceptable_count/total*100:.1f}%)"],
        ["  Weak (0.3-0.4)", f"{weak_count} ({weak_count/total*100:.1f}%)"],
        ["  Poor (0.0-0.2)", f"{poor_count} ({poor_count/total*100:.1f}%)"],
        ["", ""],
        ["Detailed Score Breakdown:", ""],
    ]

    # Add detailed score breakdown
    for score in [1.0, 0.9, 0.8, 0.7, 0.6, 0.5, 0.4, 0.3, 0.2, 0.1, 0.0]:
        count = score_counts[score]
        if count > 0:
            summary_data.append([f"  Score {score:.1f}", f"{count} ({count/total*100:.1f}%)"])

    summary_data.append(["", ""])
    summary_data.append(["By Category:", ""])

    # Add category breakdown
    categories = set(r.get("category", "") for r in results)
    for cat in sorted(categories):
        cat_results = [r for r in results if r.get("category") == cat]
        cat_avg = sum(r.get("score", 0) for r in cat_results) / len(cat_results) if cat_results else 0
        summary_data.append([f"  {cat}", f"{cat_avg:.3f} avg ({len(cat_results)} questions)"])

    # Add data quality statistics
    summary_data.append(["", ""])
    summary_data.append(["=" * 50, ""])
    summary_data.append(["DATA QUALITY ANALYSIS", ""])
    summary_data.append(["(Góc nhìn 'người chưa đọc sách')", ""])
    summary_data.append(["=" * 50, ""])

    # Count by status
    status_counts = {"OK": 0, "OK (warning)": 0, "CRITICAL": 0}
    for r in results:
        status = r.get("data_quality_status", "OK")
        if status in status_counts:
            status_counts[status] += 1

    items_ok_total = status_counts["OK"] + status_counts["OK (warning)"]
    items_critical = status_counts["CRITICAL"]

    summary_data.append(["", ""])
    summary_data.append(["PHÂN LOẠI TRẠNG THÁI:", ""])
    summary_data.append(["  ✓ OK (hoàn toàn tốt)", f"{status_counts['OK']} câu"])
    summary_data.append(["  ✓ OK (warning - vẫn ổn)", f"{status_counts['OK (warning)']} câu ({status_counts['OK (warning)']/total*100:.1f}%)"])
    summary_data.append(["  ✗ CRITICAL (cần xem lại)", f"{items_critical} câu ({items_critical/total*100:.1f}%)"])
    summary_data.append(["", ""])
    summary_data.append(["=> CÂU HỎI VẪN OK:", f"{items_ok_total} ({items_ok_total/total*100:.1f}%)"])
    summary_data.append(["=> CẦN XEM LẠI:", f"{items_critical} ({items_critical/total*100:.1f}%)"])

    # Count by issue type
    issue_type_counts = {
        "NGU_CANH": 0,
        "TRI_THUC": 0,
        "NHIEU": 0,
        "MO_HO": 0,
        "CAU_TRUC": 0,
        "TRUNG_LAP": 0,
        "KHO_HIEU": 0,
        "NHAY_CAM": 0,
    }
    for r in results:
        issues_str = r.get("data_quality_issues", "")
        for issue_type in issue_type_counts.keys():
            if f"[{issue_type}]" in issues_str:
                issue_type_counts[issue_type] += 1

    summary_data.append(["", ""])
    summary_data.append(["CHI TIẾT LOẠI LỖI:", ""])
    summary_data.append(["", ""])
    summary_data.append(["[WARNING - Câu hỏi vẫn OK]:", ""])
    summary_data.append(["  NGU_CANH (Context ngắn)", f"{issue_type_counts['NGU_CANH']} lỗi"])
    summary_data.append(["  TRI_THUC (Tri thức chung)", f"{issue_type_counts['TRI_THUC']} lỗi"])
    summary_data.append(["  CAU_TRUC (Cấu trúc)", f"{issue_type_counts['CAU_TRUC']} lỗi"])
    summary_data.append(["", ""])
    summary_data.append(["[CRITICAL - Cần sửa nội dung]:", ""])
    summary_data.append(["  TRUNG_LAP (Trùng lặp)", f"{issue_type_counts['TRUNG_LAP']} lỗi"])
    summary_data.append(["  NHIEU (Nhiễu dữ liệu)", f"{issue_type_counts['NHIEU']} lỗi"])
    summary_data.append(["  MO_HO (Mơ hồ)", f"{issue_type_counts['MO_HO']} lỗi"])
    summary_data.append(["  KHO_HIEU (Khó hiểu)", f"{issue_type_counts['KHO_HIEU']} lỗi"])
    summary_data.append(["  NHAY_CAM (Nhạy cảm)", f"{issue_type_counts['NHAY_CAM']} lỗi"])

    for row_idx, row in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif col_idx == 1 and ":" in str(value):
                cell.font = Font(bold=True)

    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 35

    # Save
    wb.save(output_path)
    print(f"\n✓ Results saved to: {output_path}")

    # Print summary
    print(f"\n{'='*60}")
    print("EVALUATION SUMMARY")
    print(f"{'='*60}")
    print(f"Model: {model_name}")
    print(f"Total questions: {total}")
    print(f"\nScore Distribution:")
    print(f"  Excellent (0.9-1.0): {excellent_count} ({excellent_count/total*100:.1f}%)")
    print(f"  Good (0.7-0.8):      {good_count} ({good_count/total*100:.1f}%)")
    print(f"  Acceptable (0.5-0.6):{acceptable_count} ({acceptable_count/total*100:.1f}%)")
    print(f"  Weak (0.3-0.4):      {weak_count} ({weak_count/total*100:.1f}%)")
    print(f"  Poor (0.0-0.2):      {poor_count} ({poor_count/total*100:.1f}%)")
    print(f"\nAverage score: {avg_score:.3f}")
    print(f"Total points:  {sum(r.get('score', 0) for r in results):.1f} / {total}")

    # Print data quality summary
    status_counts = {"OK": 0, "OK (warning)": 0, "CRITICAL": 0}
    for r in results:
        status = r.get("data_quality_status", "OK")
        if status in status_counts:
            status_counts[status] += 1

    items_ok_total = status_counts["OK"] + status_counts["OK (warning)"]
    items_critical = status_counts["CRITICAL"]

    print(f"\nData Quality:")
    print(f"  ✓ OK (warning - vẫn ổn): {status_counts['OK (warning)']} ({status_counts['OK (warning)']/total*100:.1f}%)")
    print(f"  ✗ CRITICAL (cần xem lại): {items_critical} ({items_critical/total*100:.1f}%)")
    print(f"  => Câu hỏi vẫn OK: {items_ok_total} ({items_ok_total/total*100:.1f}%)")
    print(f"{'='*60}")


def main():
    """Main function to run LLM-as-a-Judge evaluation."""
    import argparse

    parser = argparse.ArgumentParser(description='LLM-as-a-Judge Evaluation')
    parser.add_argument('--benchmark', type=str, required=False,
                        help='Path to benchmark JSON file')
    parser.add_argument('--model', type=str, default="gpt-4o-2024-11-20",
                        help='Model to evaluate')
    parser.add_argument('--output', type=str, required=False,
                        help='Output Excel file path')
    parser.add_argument('--limit', type=int, default=None,
                        help='Limit number of questions to evaluate')

    args = parser.parse_args()

    # Default paths
    base_dir = r"D:/dichdata/vietnamese-culture-eval-2"

    if not args.benchmark:
        # Try to find benchmark files
        benchmark_paths = [
            os.path.join(base_dir, "data_question_answer", "ban_sac_van_hoa_viet_nam", "culture_benchmark.json"),
            os.path.join(base_dir, "data_question_answer", "bai_giang_phap_luat_dai_cuong", "law_benchmark.json"),
        ]

        benchmark_data = []
        for path in benchmark_paths:
            benchmark_data.extend(load_benchmark(path))
    else:
        benchmark_data = load_benchmark(args.benchmark)

    if not benchmark_data:
        print("No benchmark data found. Please generate benchmarks first using generate_qa_benchmark.py")
        return

    # Apply limit if specified
    if args.limit:
        benchmark_data = benchmark_data[:args.limit]

    # Output path
    if not args.output:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        args.output = os.path.join(base_dir, f"evaluation_results_{timestamp}.xlsx")

    print("=" * 60)
    print("LLM-AS-A-JUDGE EVALUATION")
    print("=" * 60)
    print(f"Benchmark: {len(benchmark_data)} questions")
    print(f"Model: {args.model}")
    print(f"Output: {args.output}")

    # Run evaluation
    results = evaluate_benchmark(benchmark_data, args.model)

    # Export results
    export_results_to_excel(results, args.output, args.model)


if __name__ == "__main__":
    main()
