"""
Auto Validation Scorer - Tu dong danh gia 4 tieu chi chat luong

Su dung LLM de tu dong cham diem 4 tieu chi:
1. text_based: Dua hoan toan vao van ban (1/0.5/0)
2. no_temporal: Khong co thong tin thay doi theo thoi gian (1/0.5/0)
3. relevant: Lien quan den van hoa hoac phap luat Viet Nam (1/0.5/0)
4. objective: Trung lap, khach quan (1/0.5/0)
"""

import os
import json
import openai
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Azure OpenAI Configuration
client = openai.AzureOpenAI(
    api_key="a6705b22532443ee8c0cfda232e57e06",
    azure_endpoint="https://vietgpt.openai.azure.com/",
    api_version="2024-02-15-preview"
)


VALIDATION_SYSTEM_PROMPT = """
Ban la chuyen gia danh gia chat luong cau hoi benchmark ve van hoa va phap luat Viet Nam.

## NHIEM VU:
Danh gia cau hoi theo 4 tieu chi. Moi tieu chi cho diem 1, 0.5, hoac 0.

## 4 TIEU CHI DANH GIA:

### 1. text_based (Dua vao van ban)
- 1: Cau hoi va dap an dua hoan toan vao noi dung van ban goc, nguoi doc chi can doc doan van la tra loi duoc
- 0.5: Phan lon dua vao van ban, co mot chut suy luan nhe hoac kien thuc pho thong
- 0: Suy luan ngoai van ban hoac thong tin khong co trong nguon

### 2. no_temporal (Khong thay doi theo thoi gian)
- 1: Thong tin khong thay doi theo thoi gian (khai niem, nguyen tac, truyen thong...)
- 0.5: Co de cap thoi gian nhung khong anh huong den tinh dung dan cua cau hoi
- 0: Co thong tin co the loi thoi (so lieu GDP, dan so, su kien gan day, luat moi...)

### 3. relevant (Lien quan VN)
- 1: Hoan toan lien quan den van hoa hoac phap luat Viet Nam
- 0.5: Lien quan nhung co phan chung chung hoac khong dac thu VN
- 0: Khong lien quan hoac qua xa voi

### 4. objective (Khach quan)
- 1: Trung lap, khach quan, co dap an ro rang, khong y kien ca nhan
- 0.5: Co mot chut thien vi nhung khong nghiem trong
- 0: Co y kien ca nhan, thien vi ro rang, hoac cau hoi chu quan

## DINH DANG TRA LOI (JSON):
{
  "text_based": <1 | 0.5 | 0>,
  "no_temporal": <1 | 0.5 | 0>,
  "relevant": <1 | 0.5 | 0>,
  "objective": <1 | 0.5 | 0>,
  "reasoning": "<giai thich ngan gon cho cac diem tru neu co>"
}

Chi tra ve JSON, khong giai thich them.
"""

VALIDATION_USER_TEMPLATE = """
## CONTEXT (Doan van goc):
{context}

## CAU HOI:
{question}

## DAP AN:
{answer}

Hay danh gia cau hoi nay theo 4 tieu chi.
"""


def score_validation_criteria(context: str, question: str, answer: str) -> dict:
    """
    Su dung LLM de danh gia 4 tieu chi chat luong.

    Returns:
        Dict voi text_based, no_temporal, relevant, objective scores
    """
    try:
        user_prompt = VALIDATION_USER_TEMPLATE.format(
            context=context[:1500] if len(context) > 1500 else context,  # Gioi han context
            question=question,
            answer=answer
        )

        response = client.chat.completions.create(
            model="gpt-4o-2024-11-20",
            messages=[
                {"role": "system", "content": VALIDATION_SYSTEM_PROMPT},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.1,
            max_tokens=300
        )

        result = response.choices[0].message.content.strip()

        # Parse JSON
        if "```json" in result:
            result = result.split("```json")[1].split("```")[0].strip()
        elif "```" in result:
            result = result.split("```")[1].split("```")[0].strip()

        scores = json.loads(result)

        # Validate scores
        for key in ["text_based", "no_temporal", "relevant", "objective"]:
            if key not in scores:
                scores[key] = 0
            elif scores[key] not in [0, 0.5, 1]:
                # Normalize
                if scores[key] > 0.75:
                    scores[key] = 1
                elif scores[key] > 0.25:
                    scores[key] = 0.5
                else:
                    scores[key] = 0

        return scores

    except json.JSONDecodeError as e:
        return {
            "text_based": 0,
            "no_temporal": 0,
            "relevant": 0,
            "objective": 0,
            "reasoning": f"JSON parse error: {str(e)}"
        }
    except Exception as e:
        return {
            "text_based": 0,
            "no_temporal": 0,
            "relevant": 0,
            "objective": 0,
            "reasoning": f"Error: {str(e)}"
        }


def run_validation_scoring(data: list, progress_file: str = None) -> list:
    """
    Chay validation scoring cho tat ca items.
    Ho tro resume tu progress file.

    Args:
        data: List of Q&A dictionaries
        progress_file: Path to save/load progress

    Returns:
        List of Q&A dictionaries with validation scores
    """
    # Load existing progress
    processed = {}
    if progress_file and os.path.exists(progress_file):
        try:
            with open(progress_file, 'r', encoding='utf-8') as f:
                processed = json.load(f)
            print(f"  Loaded {len(processed)} existing validation scores")
        except:
            pass

    total = len(data)
    scored_count = 0

    for idx, item in enumerate(data, 1):
        item_id = item.get("id", f"Q{idx}")

        # Skip if already processed
        if item_id in processed:
            item["text_based"] = processed[item_id].get("text_based", 0)
            item["no_temporal"] = processed[item_id].get("no_temporal", 0)
            item["relevant"] = processed[item_id].get("relevant", 0)
            item["objective"] = processed[item_id].get("objective", 0)
            item["validation_reasoning"] = processed[item_id].get("reasoning", "")
            continue

        if idx % 10 == 1 or idx == 1:
            print(f"  [{idx}/{total}] Scoring validation criteria...")

        context = item.get("context", "")
        question = item.get("question", "")
        answer = item.get("answer", "")

        # Get scores
        scores = score_validation_criteria(context, question, answer)

        item["text_based"] = scores.get("text_based", 0)
        item["no_temporal"] = scores.get("no_temporal", 0)
        item["relevant"] = scores.get("relevant", 0)
        item["objective"] = scores.get("objective", 0)
        item["validation_reasoning"] = scores.get("reasoning", "")
        scored_count += 1

        # Save to processed dict
        processed[item_id] = {
            "text_based": item["text_based"],
            "no_temporal": item["no_temporal"],
            "relevant": item["relevant"],
            "objective": item["objective"],
            "reasoning": item["validation_reasoning"]
        }

        # Save progress after EACH item
        if progress_file:
            try:
                with open(progress_file, 'w', encoding='utf-8') as f:
                    json.dump(processed, f, ensure_ascii=False, indent=2)
            except Exception as e:
                print(f"    Warning: Error saving progress: {e}")

        # Print progress
        if scored_count % 20 == 0:
            print(f"    Progress: {scored_count} scored")

    print(f"  Completed: {scored_count} new scores")
    return data


def load_benchmark_data(base_dir: str) -> list:
    """Load benchmark data from JSON files."""
    all_data = []

    qa_dir = os.path.join(base_dir, "data_question_answer")

    # Culture benchmark
    culture_path = os.path.join(qa_dir, "ban_sac_van_hoa_viet_nam", "culture_benchmark.json")
    if os.path.exists(culture_path):
        with open(culture_path, 'r', encoding='utf-8') as f:
            culture_data = json.load(f)
        print(f"  Loaded {len(culture_data)} culture questions")
        all_data.extend(culture_data)

    # Law benchmark
    law_path = os.path.join(qa_dir, "bai_giang_phap_luat_dai_cuong", "law_benchmark.json")
    if os.path.exists(law_path):
        with open(law_path, 'r', encoding='utf-8') as f:
            law_data = json.load(f)
        print(f"  Loaded {len(law_data)} law questions")
        all_data.extend(law_data)

    return all_data


def load_existing_data(base_dir: str, data: list) -> list:
    """Load existing topics, GPT answers, and judge scores into data."""

    # Load topics
    topics_file = os.path.join(base_dir, "topics_progress.json")
    if os.path.exists(topics_file):
        try:
            with open(topics_file, 'r', encoding='utf-8') as f:
                topics = json.load(f)
            for item in data:
                item_id = item.get("id", "")
                if item_id in topics:
                    item["predicted_topic"] = topics[item_id]
            print(f"  Loaded {len(topics)} topic predictions")
        except Exception as e:
            print(f"  Warning: Error loading topics: {e}")

    # Load GPT web answers
    for web_file in ["chatgpt_web_answers.json", "chatgpt_web_answers_law.json"]:
        web_path = os.path.join(base_dir, web_file)
        if os.path.exists(web_path):
            try:
                with open(web_path, 'r', encoding='utf-8') as f:
                    web_answers = json.load(f)
                matched = 0
                for item in data:
                    item_id = item.get("id", "")
                    if item_id in web_answers:
                        answer_data = web_answers[item_id]
                        if isinstance(answer_data, dict):
                            item["gpt_answer"] = answer_data.get("gpt_web_answer", "")
                        else:
                            item["gpt_answer"] = str(answer_data)
                        matched += 1
                print(f"  Loaded {matched} GPT answers from {web_file}")
            except Exception as e:
                print(f"  Warning: Error loading {web_file}: {e}")

    # Load judge scores
    judge_file = os.path.join(base_dir, "judge_scores_progress.json")
    if os.path.exists(judge_file):
        try:
            with open(judge_file, 'r', encoding='utf-8') as f:
                judge_scores = json.load(f)
            for item in data:
                item_id = item.get("id", "")
                if item_id in judge_scores:
                    item["judge_score"] = judge_scores[item_id].get("score")
                    item["judge_verdict"] = judge_scores[item_id].get("verdict", "")
                    item["judge_reason"] = judge_scores[item_id].get("reason", "")
            print(f"  Loaded {len(judge_scores)} judge scores")
        except Exception as e:
            print(f"  Warning: Error loading judge scores: {e}")

    return data


def create_final_excel(data: list, output_path: str):
    """
    Tao file Excel cuoi cung voi day du cot validation va sort.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation"

    # Headers
    headers = [
        "ID",
        "Source",
        "Category",
        "Predicted Topic",
        "Context",
        "Question",
        "Answer",
        "GPT_Answer",
        "GPT Score",
        "GPT Verdict",
        "GPT Reason",
        "text_based\n(Dua vao VB)",
        "no_temporal\n(Kg TG)",
        "relevant\n(Lien quan)",
        "objective\n(Khach quan)",
        "Total Score",
        "Validation Reasoning",
        "Notes"
    ]

    col_widths = [10, 20, 10, 15, 35, 45, 50, 50, 10, 12, 40, 12, 12, 12, 12, 12, 40, 30]

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap_align = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Set column widths
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Sort data by predicted_topic
    sorted_data = sorted(data, key=lambda x: (x.get("predicted_topic", "Khac"), x.get("id", "")))

    # Topic colors
    topic_colors = {
        "Phap luat": "B4C6E7",
        "Van hoa": "F8CBAD",
        "Lich su": "FFE699",
        "Kinh te": "C6EFCE",
        "Chinh tri": "D9E1F2",
        "Xa hoi": "E2EFDA",
        "Giao duc": "FCE4D6",
        "Ton giao": "DDEBF7",
        "Nghe thuat": "FFF2CC",
        "Phong tuc tap quan": "E4DFEC",
        "Dia ly": "D0CECE",
        "Khac": "F2F2F2"
    }

    # Score colors
    score_colors = {
        1.0: "00B050",
        0.5: "FFC000",
        0.0: "FF6666",
    }

    # Write data rows
    for row_idx, item in enumerate(sorted_data, 2):
        col = 1

        # ID
        ws.cell(row=row_idx, column=col, value=item.get("id", "")).border = thin_border
        col += 1

        # Source
        ws.cell(row=row_idx, column=col, value=item.get("source", "")).border = thin_border
        col += 1

        # Category
        ws.cell(row=row_idx, column=col, value=item.get("category", "")).border = thin_border
        col += 1

        # Predicted Topic
        topic = item.get("predicted_topic", "Khac")
        topic_cell = ws.cell(row=row_idx, column=col, value=topic)
        topic_cell.alignment = center_align
        topic_cell.border = thin_border
        if topic in topic_colors:
            topic_cell.fill = PatternFill(start_color=topic_colors[topic], end_color=topic_colors[topic], fill_type="solid")
        col += 1

        # Context
        ctx_cell = ws.cell(row=row_idx, column=col, value=item.get("context", ""))
        ctx_cell.alignment = wrap_align
        ctx_cell.border = thin_border
        col += 1

        # Question
        q_cell = ws.cell(row=row_idx, column=col, value=item.get("question", ""))
        q_cell.alignment = wrap_align
        q_cell.border = thin_border
        col += 1

        # Answer
        a_cell = ws.cell(row=row_idx, column=col, value=item.get("answer", ""))
        a_cell.alignment = wrap_align
        a_cell.border = thin_border
        col += 1

        # GPT Answer
        gpt_cell = ws.cell(row=row_idx, column=col, value=item.get("gpt_answer", ""))
        gpt_cell.alignment = wrap_align
        gpt_cell.border = thin_border
        if item.get("gpt_answer"):
            gpt_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        col += 1

        # GPT Score
        judge_score = item.get("judge_score")
        score_cell = ws.cell(row=row_idx, column=col, value=judge_score if judge_score is not None else "")
        score_cell.alignment = center_align
        score_cell.border = thin_border
        col += 1

        # GPT Verdict
        ws.cell(row=row_idx, column=col, value=item.get("judge_verdict", "")).border = thin_border
        col += 1

        # GPT Reason
        reason_cell = ws.cell(row=row_idx, column=col, value=item.get("judge_reason", ""))
        reason_cell.alignment = wrap_align
        reason_cell.border = thin_border
        col += 1

        # Validation scores (4 columns)
        validation_start_col = col
        for score_key in ["text_based", "no_temporal", "relevant", "objective"]:
            score_val = item.get(score_key)
            v_cell = ws.cell(row=row_idx, column=col, value=score_val if score_val is not None else "")
            v_cell.alignment = center_align
            v_cell.border = thin_border
            # Color by score
            if score_val in score_colors:
                v_cell.fill = PatternFill(start_color=score_colors[score_val], end_color=score_colors[score_val], fill_type="solid")
            col += 1

        # Total Score (formula)
        validation_end_col = col - 1
        total_cell = ws.cell(
            row=row_idx,
            column=col,
            value=f"=SUM({get_column_letter(validation_start_col)}{row_idx}:{get_column_letter(validation_end_col)}{row_idx})"
        )
        total_cell.alignment = center_align
        total_cell.border = thin_border
        col += 1

        # Validation Reasoning
        reason_cell = ws.cell(row=row_idx, column=col, value=item.get("validation_reasoning", ""))
        reason_cell.alignment = wrap_align
        reason_cell.border = thin_border
        col += 1

        # Notes (empty)
        notes_cell = ws.cell(row=row_idx, column=col, value="")
        notes_cell.alignment = wrap_align
        notes_cell.border = thin_border

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto filter
    last_col = get_column_letter(len(headers))
    last_row = len(sorted_data) + 1
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Add summary sheet
    ws_summary = wb.create_sheet("Tong ket")

    # Calculate statistics
    total_items = len(data)

    # Count by topic
    topic_counts = {}
    for item in data:
        topic = item.get("predicted_topic", "Khac")
        topic_counts[topic] = topic_counts.get(topic, 0) + 1

    # Average validation scores
    def safe_avg(scores):
        valid = [s for s in scores if s is not None]
        return sum(valid) / len(valid) if valid else 0

    avg_text_based = safe_avg([d.get("text_based") for d in data])
    avg_no_temporal = safe_avg([d.get("no_temporal") for d in data])
    avg_relevant = safe_avg([d.get("relevant") for d in data])
    avg_objective = safe_avg([d.get("objective") for d in data])
    avg_total = avg_text_based + avg_no_temporal + avg_relevant + avg_objective

    # Judge score stats
    judge_scores = [d.get("judge_score") for d in data if d.get("judge_score") is not None]
    avg_judge = sum(judge_scores) / len(judge_scores) if judge_scores else 0

    # Write summary
    summary_data = [
        ["TONG KET BENCHMARK", ""],
        ["", ""],
        ["Tong so cau hoi", total_items],
        ["Co GPT answer", sum(1 for d in data if d.get("gpt_answer"))],
        ["Co Judge score", len(judge_scores)],
        ["", ""],
        ["DIEM DANH GIA TRUNG BINH", ""],
        ["text_based", f"{avg_text_based:.2f}"],
        ["no_temporal", f"{avg_no_temporal:.2f}"],
        ["relevant", f"{avg_relevant:.2f}"],
        ["objective", f"{avg_objective:.2f}"],
        ["Total (max 4)", f"{avg_total:.2f}"],
        ["", ""],
        ["GPT Judge Score TB", f"{avg_judge:.3f}"],
        ["", ""],
        ["PHAN BO CHU DE", ""],
    ]

    for topic, count in sorted(topic_counts.items(), key=lambda x: -x[1]):
        summary_data.append([f"  {topic}", f"{count} ({count/total_items*100:.1f}%)"])

    for row_idx, (col1, col2) in enumerate(summary_data, 1):
        ws_summary.cell(row=row_idx, column=1, value=col1)
        ws_summary.cell(row=row_idx, column=2, value=col2)
        if "TONG KET" in str(col1) or "DIEM DANH GIA" in str(col1) or "PHAN BO" in str(col1):
            ws_summary.cell(row=row_idx, column=1).font = Font(bold=True)

    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20

    # Save
    wb.save(output_path)
    print(f"\n Excel saved to: {output_path}")
    print(f"  - Total: {total_items} questions")
    print(f"  - Sorted by: predicted_topic")
    print(f"  - Sheets: Validation, Tong ket")


def main():
    """Main function."""
    import argparse

    parser = argparse.ArgumentParser(description='Auto Validation Scorer')
    parser.add_argument('--no-score', action='store_true',
                        help='Skip validation scoring (use existing progress)')
    parser.add_argument('--limit', type=int, default=None,
                        help='Limit number of questions')
    args = parser.parse_args()

    base_dir = r"D:\dichdata\vietnamese-culture-eval-2"
    output_path = os.path.join(base_dir, "vietnamese_benchmark_final.xlsx")
    validation_progress_file = os.path.join(base_dir, "validation_scores_progress.json")

    print("=" * 60)
    print("AUTO VALIDATION SCORER")
    print("=" * 60)

    # Load data
    print("\n[1/4] Loading benchmark data...")
    data = load_benchmark_data(base_dir)

    if not data:
        print("  No data found!")
        return

    if args.limit:
        data = data[:args.limit]
        print(f"  Limited to {args.limit} questions")

    # Load existing data
    print("\n[2/4] Loading existing data (topics, GPT answers, judge scores)...")
    data = load_existing_data(base_dir, data)

    # Run validation scoring
    if not args.no_score:
        print(f"\n[3/4] Running validation scoring for {len(data)} questions...")
        print("  (Progress saved after each item. You can stop and resume.)")
        data = run_validation_scoring(data, validation_progress_file)
    else:
        print("\n[3/4] Loading existing validation scores...")
        if os.path.exists(validation_progress_file):
            with open(validation_progress_file, 'r', encoding='utf-8') as f:
                processed = json.load(f)
            for item in data:
                item_id = item.get("id", "")
                if item_id in processed:
                    item["text_based"] = processed[item_id].get("text_based", 0)
                    item["no_temporal"] = processed[item_id].get("no_temporal", 0)
                    item["relevant"] = processed[item_id].get("relevant", 0)
                    item["objective"] = processed[item_id].get("objective", 0)
                    item["validation_reasoning"] = processed[item_id].get("reasoning", "")
            print(f"  Loaded {len(processed)} existing scores")
        else:
            print("  No existing scores found")

    # Create Excel
    print(f"\n[4/4] Creating final Excel file...")
    create_final_excel(data, output_path)

    print("\n" + "=" * 60)
    print("COMPLETED!")
    print("=" * 60)
    print(f"\nOutput: {output_path}")


if __name__ == "__main__":
    main()
